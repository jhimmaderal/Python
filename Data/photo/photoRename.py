from openpyxl import load_workbook as lw
import os
import re


def converNote():
    photoName = []
    source = "D:\photos"
    # with open("barcode.txt", "w") as txt_file:
    for photo in os.listdir(source):
        if os.path.isfile(os.path.join(source, photo)):
            removeExt = photo[: (len(photo) - 4)]
            # print(removeExt)
            # txt_file.write("".join(removeExt) + "\n")
            photoName.append(removeExt)
    return photoName


def createDict(photoName):
    dictPhoto = {}
    path = "list.xlsx"
    xlsFile = lw(path)
    shtFile = xlsFile.active
    rowFile = shtFile.max_row
    colFile = shtFile.max_column

    for item in range(1, rowFile):
        fpCode = shtFile.cell(row=item, column=1)
        fpVal = fpCode.value
        itemCode = shtFile.cell(row=item, column=2)
        itemVal = itemCode.value
        for photo in photoName:
            if isinstance(photo, str) and re.search(str(fpVal), photo):
                filename = {str(photo): str(itemVal)}
                dictPhoto.update(filename)
    # print(dictPhoto)
    return dictPhoto


def renameFile(dictPhoto):
    source = "D:\photos"

    for key, value in dict(dictPhoto).items():
        try:
            filename = key + ".jpg"
            # print(value)
            if value == "#N/A":
                print(key, "No item Code")
            else:
                os.rename(
                    os.path.join(source, filename), os.path.join(source, value + ".jpg")
                )
        except FileExistsError:
            print(value, "Filename already exist")


photoName = converNote()
dictPhoto = createDict(photoName)
renameFile(dictPhoto)
