import pyautogui as py
import time
from openpyxl import load_workbook as lw

def openFile():
  itemList =[]
  filePath = "uom.xlsx"
  exlFile = lw(filePath)
  shtFile = exlFile.active
  rowFile = shtFile.max_row
    
  for item in range(2, rowFile + 1):
    for col in range(1,4):
      itemCell = shtFile.cell(row=item, column = col) # loop from column
      itemVal = itemCell.value
      itemList.append(itemVal)
  
  print(itemList)
  return itemList
  
     
     
def addUom(itemList):
  step = 0
  py.click(30,15)
  for item in itemList:
    match step:
      case 0:
        time.sleep(2)
        py.press('f6')
        py.press('f5')
        py.press('tab',presses=2)
        py.write(str(item))
        py.press('enter',presses=2)
        step +=1
      case 1:
        py.press('f8')
        py.press('tab',presses=15)
        py.press('right',presses=6)
        py.press('tab',presses=9)
        py.write(str(item))
        py.press('tab')
        step +=1
      case 2:
        py.press('.')
        py.press('del', presses=3)
        py.press('left', presses=4)
        py.write(str(item))
        py.press('f10')
        step = 0

def runApp():   
  run = input("Do you want to start UOM App ? Y/N: ").lower()
  if run ==  'y':
    itemList = openFile()
    addUom(itemList)
  else:
    print("Ok !")
    exit()