import pyautogui as pg
import time
from openpyxl import load_workbook as lw

file = "price change.xlsx"
items = []
xslFile = lw(file)
shtFile = xslFile.active
rowFile = shtFile.max_row  # count total row
colFile = shtFile.max_column  # count total column

counter = 0

if counter == 0:
  for item in range(3,rowFile + 1): # loop to A3 to last item
    
    codeCell = shtFile.cell(row=item, column = 1)
    codeVal = codeCell.value
    items.append(codeVal)
    
    costCell = shtFile.cell(row=item, column = 3)
    costVal = costCell.value
    items.append(costVal)
    
    retailCell = shtFile.cell(row=item, column = 4)
    retailVal = retailCell.value   
    items.append(retailVal)
      
print(items)

def priceChange():
  time.sleep(3)
  for i in items:
    pg.write(str(i))
    pg.press("enter")
    
priceChange()








