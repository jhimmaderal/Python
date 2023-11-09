import pyautogui as pg
import time 
from openpyxl import load_workbook as lw

itemCost = []

def OpenFile():
  startRow = 3
  print("1. Create New")
  print("2. Continue Last Session")
  createNew = input("Select Item: ")
  if createNew == "1":
    startRow = 3
  elif createNew == "2":
    startRow = int(input("Select Row: "))
  else:
    print("Not in the list")
  
  file = "price change.xlsx"
  xslFile = lw(file)
  shtFile = xslFile.active #change Sheet ["sheetName"]
  rowFile = shtFile.max_row  # count total row
  colFile = shtFile.max_column  # count total column
   
  for item in range(startRow,rowFile + 1): # loop to A3 to last item
    codeCell = shtFile.cell(row=item, column = 1)
    codeVal = codeCell.value
    itemCost.append(codeVal)
    costCell = shtFile.cell(row=item, column = 3)
    costVal = costCell.value
    itemCost.append(costVal)
    landCell = shtFile.cell(row=item, column = 4)
    landVal = landCell.value
    itemCost.append(landVal)     

def costChange():
  steps = 0
  
  # Start
  pg.click(58,17, button='left')
  for item in itemCost:
    if steps == 0:
      if (len(str(item))) == 6:
        time.sleep(3)
        pg.press('f6')
        pg.press('enter')
        pg.write(str(item))
        pg.press('enter')
        pg.press('f8')
        steps = steps + 1
      else:
        time.sleep(3)
        pg.press('f6')
        pg.press('f5')
        pg.press('tab')
        pg.write(str(item))
        pg.press('enter', presses = 2)
        pg.press('f8')
        steps = steps + 1
      
    elif steps == 1:
      pg.press('tab', presses=18)
      pg.write(".")
      pg.press('delete', presses=3)
      pg.press('left', presses= 4)
      pg.write(str(item))
      pg.press('enter')
      steps = steps + 1
          
    elif steps == 2:
      pg.write(".")
      pg.press('delete', presses=3)
      pg.press('left', presses= 4)
      pg.write(str(item))
      pg.press('f10')
      steps = 0
      
def runCostChange():
  createNew = input("Start New Cost Change? Y/N? ").lower()
  
  if createNew == "y":
    OpenFile()
    print(itemCost)
    print("Price Change - Running.....")
    costChange()
  elif createNew == "n":
    exit()
  else:
    print("Not in the list")
    
