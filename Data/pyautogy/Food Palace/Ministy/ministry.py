import pyautogui as pg
import time
from openpyxl import load_workbook as lw

def ministry():
    print('Running...')
    pg.click(58,17, button='left')
    pg.press("tab", presses=26)

    for item in lists:
        pg.write(str(item))
        pg.press("tab", presses=3)
    print('Done :D')
print("Run the app once in the Ministry Stock Adjustment")
runApp = input("Do you want to run the script? Y/N:").lower()
if runApp == "y": 
    lists = []
    path = "Z:\Jim\MOI\Jhim PSMS MOI Updating.xlsx"
    xslFile = lw(path)
    shtFile = xslFile['MOI Data'] # change sheet
    rowFile = shtFile.max_row  # count total row
    colFile = shtFile.max_column  # count total column
    for i in range(4, rowFile + 1):
        cllFile = shtFile.cell(row=i, column=3)
        valFile = round(cllFile.value)
        lists.append(valFile)
        
    ministry()
else:
    print("Better luck next time :D ")
        




