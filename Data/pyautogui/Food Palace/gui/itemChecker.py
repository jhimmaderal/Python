import pyautogui as pg
import time
from openpyxl import load_workbook as lw


def checkFile():
    items = []
    startRow = 2

    file = "resources/itemChecker/itemMissing.xlsx"
    xslFile = lw(file)
    shtFile = xslFile["Items"]  # change Sheet ["sheetName"]
    rowFile = shtFile.max_row  # count total row
    colFile = shtFile.max_column  # count total column

    for list in range(startRow, rowFile + 1):  # loop to A3 to last item
        for col in range(1, colFile + 1):
            itemCell = shtFile.cell(row=list, column=col)  # loop from column
            itemVal = itemCell.value
            items.append(itemVal)

    print(f"Total Items : {rowFile}\n")
    return items


def checkItem(items):
    step = 0
    time.sleep(1)
    pg.click(197, 19, button="left")  # Clicking the App

    for item in items:
        if step == 0:
            pg.press("f6")
            pg.write(str(item))
            pg.press("enter")
            step = step + 1
        if step == 1:
            time.sleep(0.2)
            try:
                error = pg.locateOnScreen(
                    "resources/itemChecker/error.png", confidence=0.9
                )
                pg.press("enter", presses=2, interval=0.2)
                if len(error) > 0:
                    print(item)
                step = 0
            except:
                step = 0


def startApp():
    print(
        """
          1. Open System App Press Window + Left
          2. Run the script
          3. Wait until it finish
          4. Copy List from the CMD to Excel"""
    )
    start = input("Do you want to run the App? Y/N: ")
    if "y" or "Y" in start:
        items = checkFile()
        checkItem(items)
    else:
        exit


startApp()
