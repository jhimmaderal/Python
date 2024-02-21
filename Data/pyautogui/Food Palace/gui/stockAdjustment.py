import pyautogui as pg
import time
from openpyxl import load_workbook as lw

def stockAdjustment():
    reason = str('ITEM SPLITTED')
       
    def checkFile():
        items = []
        startRow = 3

        file = "fpAppFile.xlsx"
        xslFile = lw(file)
        shtFile = xslFile["Adjustment"] #change Sheet ["sheetName"]
        rowFile = shtFile.max_row  # count total row
        colFile = shtFile.max_column  # count total column

        for list in range(startRow, rowFile + 1):  # loop to A3 to last item
            for col in range(1, colFile + 1):
                itemCell = shtFile.cell(row=list, column=col)  # loop from column
                itemVal = itemCell.value
                items.append(itemVal)

        print(f"Total Items : {rowFile - 1}\n")
        return items

    def checkItem(items):
        step = 0
        time.sleep(1)
        pg.click(197, 19, button="left")  # Clicking the App

        pg.press("f9")
        pg.press('tab')
        pg.write(str(date))
        pg.press('tab',presses=3)
        pg.write(str(reason))

        for item in items:
            match step:
                case 0:
                    pg.press("f11")
                    pg.write("@" + str(item))
                    step = step + 1

                case 1:
                    pg.press('tab')
                    pg.write(str(item))
                    step = step + 1
                    

                case 2:
                    pg.press('tab')
                    pg.write(str(item))
                    pg.press('enter')
                    step = 0
            

    print(
        """
          1. Open System App Press Window + Left
          2. Run the script
          3. Wait until it finish
          4. Copy List from the CMD to Excel"""
    )
    start = input("Do you want to run the App? Y/N: ")
    if "y" or "Y" in start:
        date =  input(str('Enter Date: '))
        items = checkFile()
        print(items)
        validation = input("Validate the items? Y/N: ")
        if validation == "y" or "Y":
            checkItem(items)
            amount1 = input('Enter Amount1: ')
            amount2 = input('Enter Amount2: ')
        else:
             start()
    else:
        exit()


