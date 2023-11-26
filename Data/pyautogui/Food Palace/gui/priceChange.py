import pyautogui as pg
import time
from openpyxl import load_workbook as lw

def appPriceChange():
    
    supplier = input("Enter Supplier Code: ").upper()
    print("1. Create New")
    print("2. Continue Last Session")   
    createNew = int(input("Select Item: "))
    if createNew == 1:
        startRow = 3
    elif createNew == 2:
        startRow = int(input("Select Row:  Default is 3: "))
    else:
        createNew = int(input("Not in the list ! Select Item: "))
        
    print("Price Change - Running.....") 

    file = "fpAppFile.xlsx"
    itemPrice = []
    xslFile = lw(file)
    shtFile = xslFile["Price Change"] #change Sheet ["sheetName"]
    rowFile = shtFile.max_row  # count total row
    colFile = shtFile.max_column  # count total column

    for item in range(startRow,rowFile + 1): # loop to A3 to last item
        codeCell = shtFile.cell(row=item, column = 1)
        codeVal = codeCell.value
        itemPrice.append(codeVal)
        retailCell = shtFile.cell(row=item, column = 4)
        retailVal = retailCell.value
        itemPrice.append(retailVal)
        
    def goToPriceChange():
        pg.click(58,17, button='left')
        pg.keyDown('alt')
        pg.press('r',presses=2)
        pg.keyUp('alt')
        pg.press('enter')
        pg.press('p', presses= 2)
        pg.press('enter')   
    def createChange():
        pg.press("f9")
        pg.write("HO")
        pg.press('tab', presses= 5)
        pg.typewrite(supplier)
        pg.press('tab', presses= 13)
        pg.typewrite("PRICE CHANGE")
    def createLines():
        pg.press("f11")
        steps = 0
        for item in itemPrice:
            if steps == 0:
                if len(str(item)) == 6:
                    pg.write("@" + str(item))
                    pg.press('enter')
                    steps = steps + 1
                else:
                    pg.write("@" + str(item))
                    pg.press('enter')
                    steps = steps + 1
            else:
                pg.press("tab", presses=6)
                pg.write(".")
                pg.press("delete", presses=2)
                pg.press("left",presses=3)
                pg.write(str(item))
                pg.press("enter")
                time.sleep(0.5)
                steps = steps - 1 
        itemPrice = []
        
    try:
        if createNew == 1:
            pg.click(58,17, button='left')
            time.sleep(2)
            goToPriceChange()
            createChange()
            createLines()
            print(itemPrice)
            
        elif createNew == 2:
            pg.click(58,17, button='left')
            time.sleep(2)
            createLines()
            print(itemPrice)

        else:
            print("Not in the list! ")
            createNew = input("Select Item: ")
    except ValueError:
        createNew = input("Select Item: ")
        
    appPriceChange()




