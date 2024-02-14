import pyautogui as pg
import time
from openpyxl import load_workbook as lw

def checkFile():
    items = []
    startRow = 2

    file = "fpAppFile.xlsx"
    xslFile = lw(file)
    shtFile = xslFile["Supplier Change"] #change Sheet ["sheetName"]
    rowFile = shtFile.max_row  # count total row
    colFile = shtFile.max_column  # count total column

    for list in range(startRow, rowFile + 1):  # loop to A3 to last item
        for col in range(1, colFile + 1):
            itemCell = shtFile.cell(row=list, column=col)  # loop from column
            itemVal = itemCell.value
            items.append(itemVal)

    print(f"Total Items : {rowFile - 1}\n")
    return items


def checkItem(items):
    suplierCode = "SP0796"
    step = 0
    pg.click(197, 19, button="left")  # Clicking the App

    for item in items:
        if step == 0:
            time.sleep(1)
            pg.press("f6")
            pg.press("f5")
            pg.press("tab", presses=2)
            pg.write(str(item))
            pg.press("enter",presses=2, interval=.5)
            step = step + 1
        if step == 1:
            pg.press("f8")
            time.sleep(1)
            try:
              priceCode = pg.locateCenterOnScreen(
                    "resources/SupplierChange/priceCode.bmp", confidence=0.9
              )
              if len(priceCode) > 1:
                pg.press("f8")
                pg.leftClick(priceCode)
                pg.hotkey("ctrl", "a")
                pg.write(str(suplierCode))
                step = step + 1
              else:
                pg.press("f8")
                step = step + 1
            except:
              pg.press("f8")
              step = step + 1
              
        if step == 2:
            time.sleep(0.5)
            reference = pg.locateCenterOnScreen(
                "resources/SupplierChange/reference.bmp", confidence=0.9
            )
            pg.leftClick(reference)
            step = step + 1
        if step ==3:
            time.sleep(0.5)
            try:
              supCode = pg.locateCenterOnScreen(
                  "resources/SupplierChange/supcode.png", confidence=0.9
              )
              if len(supCode) > 1:
                pg.leftClick(supCode)
                step = step + 1
              else:
                pg.press("f10")
                step = 0
            except:
              pg.press("f10")
              step = 0
        if step == 4:
            pg.write(suplierCode)
            pg.press("f10")
            step = 0

def startApp():
    print(
        """
          1. Open System App Press Window + Left
          2. Run the script
          3. Wait until it finish
          4. Copy List from the CMD to Excel"""

    )
    reference = pg.locateCenterOnScreen(
                "resources/SupplierChange/reference.bmp", confidence=0.85
            )
    print(reference)
    
    start = input("Do you want to run the App? Y/N: ")
    if "y" or "Y" in start:
        items = checkFile()
        checkItem(items)
    else:
        exit


startApp()