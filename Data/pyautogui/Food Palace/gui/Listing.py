import pyautogui as pg
import time
from openpyxl import load_workbook as lw


def appListing():
    itemList = []
    startRow = 4

    lastSession = input("Continue Last Session? Y/N: ").lower()

    if lastSession == "y":
        startSelect = int(input("Input Start Row: "))
        if startSelect <= 4:
            startRow = 4
        else:
            startRow = startSelect
    else:
        startRow = 4

    file = "fpAppfile.xlsx"
    xslFile = lw(file)
    shtFile = xslFile["Listing"]  # change Sheet ["sheetName"]
    rowFile = shtFile.max_row  # count total row
    colFile = shtFile.max_column  # count total column

    supplierName = input(str("Type Supplier Code: \n"))

    for list in range(startRow, rowFile + 1):  # loop to A3 to last item
        for col in range(2, colFile + 1):
            itemCell = shtFile.cell(row=list, column=col)  # loop from column
            itemVal = itemCell.value
            itemList.append(itemVal)

    for item in itemList:
        if len(str(item)) >= 12 and len(str(item)) <= 13:
            print(f"\n{item}")
        elif len(str(item)) >= 20:
            print(f"{item}")
        else:
            pass

    def createLine():
        steps = 0
        time.sleep(2)
        pg.click(197, 19, button="left")  # Clicking the App

        for item in itemList:
            if steps == 0:  # Barcode
                pg.press("f9")
                time.sleep(2)
                if len(str(item)) < 4:
                    pg.press("tab")
                    steps = steps + 1
                else:
                    pg.write(str(item).upper())
                    pg.press("tab")
                    steps = steps + 1

            elif steps == 1:  # Group
                pg.write(str(item).upper())
                pg.press("tab")
                steps = steps + 1

            elif steps == 2:  # SS Group
                pg.write(str(item).upper())
                pg.press("tab")
                steps = steps + 1

            elif steps == 3:  # SSS Group
                pg.hotkey("alt", "down")
                pg.write(str(item).upper())
                pg.press("tab")
                steps = steps + 1

            elif steps == 4:  # Brand
                pg.write(str(item).upper())
                steps = steps + 1
                pg.press("tab")

            elif steps == 5:  # Category
                pg.press("delete", presses=4)
                pg.write(str(item).upper())
                pg.press("tab")
                steps = steps + 1

            elif steps == 6:  # Description
                pg.write(str(item).upper())
                pg.press("tab", presses=2)
                steps = steps + 1

            elif steps == 7:  # Small Description
                pg.write(str(item).upper())
                pg.press("tab", presses=2)
                steps = steps + 1

            elif steps == 8:  # Item Code
                pg.write(str(item).upper())
                pg.press("tab")
                steps = steps + 1

            elif steps == 9:  # UOM
                pg.write(str(item).upper())
                pg.press("tab")
                steps = steps + 1

            elif steps == 10:  # Base Packing
                pg.press("delete", presses=4)
                pg.write(str(item).upper())
                pg.press("tab")
                steps = steps + 1

            elif steps == 11:  # Major Packing
                pg.write(str(item).upper())
                pg.press("tab")
                steps = steps + 1

            elif steps == 12:  # Major Packing Qty
                if item > 1:
                    pg.write(".")
                    pg.press("left", 2)
                    pg.write(str(item).upper())
                    pg.press("tab")
                    steps = steps + 1
                else:
                    steps = steps + 1

            elif steps == 13:  # Major Packing Name
                pg.write(str(item).upper())
                pg.press("enter")
                steps = steps + 1

            elif steps == 14:  # Supplier & Pack ID
                pg.write(str(supplierName).upper())
                pg.press("enter")
                pg.write(str(item).upper())
                pg.press("enter")
                steps = steps + 1

            elif steps == 15:  # Cost
                pg.write(str(item).upper())
                pg.press("tab")
                steps = steps + 1

            elif steps == 16:  # Landing Cost
                pg.write(str(item).upper())
                pg.press("tab")
                steps = steps + 1

            elif steps == 17:  # Retail
                pg.write(str(item).upper())
                pg.press("f10")
                time.sleep(2)
                pg.press("f8")
                # pg.press("tab",presses=16) # 42
                pg.hotkey("shift", "tab")
                pg.press("right", presses=7)
                pg.press("tab", presses=3)
                pg.press("space", presses=2)
                pg.press("f10")
                steps = 0

        print("Done !")

    def runApp():
        print(f"Extracting Files....\n")
        start = input("Confirm Items? Y/N: ").lower()
        if start == "y":
            createLine()
        else:
            start = input("Confirm Items? Y/N: ").lower()

    runApp()
