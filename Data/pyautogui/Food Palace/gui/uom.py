import pyautogui as py
import time
from openpyxl import load_workbook as lw


def appUom():
    def openFile():
        itemList = []
        filePath = "fpAppFile.xlsx"
        exlFile = lw(filePath)
        shtFile = exlFile["UOM"]
        rowFile = shtFile.max_row

        for item in range(3, rowFile + 1):
            for col in range(1, 4):
                itemCell = shtFile.cell(row=item, column=col)  # loop from column
                itemVal = itemCell.value
                itemList.append(itemVal)

        itemCount = 0
        for item in itemList:
            print(f"{item}")
            itemCount += 1
            breakText = itemCount % 3
            if breakText == 0:
                print(f"")

        return itemList

    def addUom(itemList):
        step = 0
        py.click(30, 15)
        for item in itemList:
            match step:
                case 0:
                    time.sleep(2)
                    py.press("f6")
                    py.press("f5")
                    py.press("tab", presses=2)
                    py.write(str(item))
                    py.press("enter", presses=2)
                    step += 1

                case 1:
                    py.press("f8")
                    py.press("tab", presses=15)
                    py.press("right", presses=6)
                    py.press("tab", presses=9)
                    py.write(str(item))
                    py.press("tab")
                    step += 1

                case 2:
                    py.press(".")
                    py.press("del", presses=3)
                    py.press("left", presses=4)
                    py.write(str(item))
                    py.press("f10")
                    step = 0

        itemList = []

    def runApp():
        print(f"Extracting Files....\n")
        itemList = openFile()
        start = input("Confirm Items? Y/N: ").lower()
        if start == "y":
            addUom(itemList)
        else:
            runApp

    runApp()


if __name__ == "__main__":
    appUom()
