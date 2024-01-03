import pyautogui as pg
import time
from openpyxl import load_workbook as lw


def appMinistry():
    def ministry():
        print("Running...")
        pg.click(58, 17, button="left")
        pg.press("tab", presses=25)

        for item in lists:
            pg.write(str(item))
            pg.press("tab", presses=3)
        print("Done :D")

    print("Run the app once in the Ministry Stock Adjustment\n")
    runApp = input("Do you want to run the script? Y/N:").lower()

    if runApp == "y":
        try:
            lists = []
            path = "Z:\Jim\MOI\Jhim PSMS MOI Updating.xlsx"
            xslFile = lw(path)
            shtFile = xslFile["MOI Data"]  # change sheet
            rowFile = shtFile.max_row  # count total row
            colFile = shtFile.max_column  # count total column
            for i in range(4, rowFile + 1):
                cllFile = shtFile.cell(row=i, column=3)
                valFile = round(cllFile.value)
                lists.append(valFile)

            print(f"Extracting Files....\n")
            itemCount = 0
            for item in lists:
                print(f"{item}")
                itemCount += 1
                breakText = itemCount % 3
            if breakText == 0:
                print(f"")

            start = input("Confirm Items? Y/N: ").lower()
            if start == "y":
                ministry()
            else:
                start = input("Confirm Items? Y/N: ").lower()
        except:
            print("Please Validate the file again")
            lists = []
            runApp = input("Do you want to run the script? Y/N:").lower()
    else:
        print("Better luck next time :D ")
